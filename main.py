from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Request
from starlette.middleware.base import BaseHTTPMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse, RedirectResponse, StreamingResponse
import urllib.parse
from pydantic import BaseModel
from typing import Optional
import sqlite3
from pathlib import Path
import os
import json
import urllib.request
import json
import tempfile
import httpx
import hashlib
import hmac
import html as _html
import time
import base64
import secrets
import re
import datetime
import zipfile
import io
import stripe

stripe.api_key = os.environ.get("STRIPE_SECRET_KEY", "")

app = FastAPI(title="BNI Manager")
BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "data" / "bni.db"
GOOGLE_REDIRECT_URI = "https://gaiaarts.org/bni/api/auth/google/callback"


def get_db():
    DB_PATH.parent.mkdir(exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            display_name TEXT DEFAULT '',
            pw_hash TEXT NOT NULL,
            pw_salt TEXT NOT NULL,
            profile_data TEXT DEFAULT '{}',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            cancelled_at TEXT DEFAULT NULL
        );
        CREATE TABLE IF NOT EXISTS contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER DEFAULT 1,
            name TEXT NOT NULL,
            reading TEXT DEFAULT '',
            company TEXT DEFAULT '',
            chapter TEXT DEFAULT '',
            category TEXT DEFAULT '',
            last_meeting_date TEXT DEFAULT '',
            business_description TEXT DEFAULT '',
            area TEXT DEFAULT '',
            birthplace TEXT DEFAULT '',
            residence TEXT DEFAULT '',
            spouse TEXT DEFAULT '',
            family TEXT DEFAULT '',
            previous_jobs TEXT DEFAULT '',
            hobbies TEXT DEFAULT '',
            experience_years TEXT DEFAULT '',
            success_key TEXT DEFAULT '',
            selling_points TEXT DEFAULT '',
            target_customers TEXT DEFAULT '',
            referral_intro TEXT DEFAULT '',
            request TEXT DEFAULT '',
            goals TEXT DEFAULT '',
            accomplishments TEXT DEFAULT '',
            interests TEXT DEFAULT '',
            networks TEXT DEFAULT '',
            skills TEXT DEFAULT '',
            introduction TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS memos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contact_id INTEGER NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS reminders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contact_id INTEGER NOT NULL,
            remind_date TEXT NOT NULL,
            message TEXT DEFAULT '',
            done INTEGER DEFAULT 0,
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS google_tokens (
            user_id INTEGER PRIMARY KEY,
            access_token TEXT,
            refresh_token TEXT,
            connected_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS referrals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            contact_id INTEGER NOT NULL,
            direction TEXT NOT NULL DEFAULT 'given',
            date TEXT DEFAULT '',
            description TEXT DEFAULT '',
            result TEXT DEFAULT '進行中',
            amount INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS one_on_ones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            contact_id INTEGER,
            contact_name TEXT DEFAULT '',
            meeting_date TEXT DEFAULT (date('now','localtime')),
            duration_minutes REAL DEFAULT 0,
            transcript TEXT DEFAULT '',
            summary TEXT DEFAULT '',
            gains_goals TEXT DEFAULT '',
            gains_accomplishments TEXT DEFAULT '',
            gains_interests TEXT DEFAULT '',
            gains_networks TEXT DEFAULT '',
            gains_skills TEXT DEFAULT '',
            referral_hints TEXT DEFAULT '',
            follow_up TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
    """)
    # 既存テーブルへのカラム追加（マイグレーション）
    for sql in [
        "ALTER TABLE contacts ADD COLUMN user_id INTEGER DEFAULT 1",
        "ALTER TABLE one_on_ones ADD COLUMN follow_up TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN auth_type TEXT DEFAULT 'password'",
        "ALTER TABLE users ADD COLUMN email TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN plan TEXT DEFAULT 'free'",
        "ALTER TABLE users ADD COLUMN stripe_customer_id TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN plan_expires TEXT DEFAULT NULL",
    ]:
        try: conn.execute(sql)
        except: pass
    conn.commit()
    conn.close()


init_db()

# ── 認証 ──────────────────────────────────────────────────
active_sessions: dict = {}  # token -> user_id
oauth_states: dict = {}     # state -> user_id (Google OAuth用)
session_created: dict = {}  # token -> created_at
SESSION_TTL = 86400 * 30   # 30日
used_sso_tokens: set = set()  # JTIリプレイ攻撃防止

def hash_pw(password: str, salt: str = None):
    if salt is None:
        salt = secrets.token_hex(16)
    h = hashlib.sha256(f"{salt}{password}".encode()).hexdigest()
    return h, salt

def verify_pw(password: str, hashed: str, salt: str) -> bool:
    return hash_pw(password, salt)[0] == hashed

def init_default_user():
    conn = get_db()
    row = conn.execute("SELECT id FROM users LIMIT 1").fetchone()
    if not row:
        h, s = hash_pw('bni2024')
        conn.execute("INSERT INTO users (username,display_name,pw_hash,pw_salt) VALUES (?,?,?,?)",
                     ('admin', '管理者', h, s))
        conn.commit()
        # 既存コンタクトをadminに割り当て
        conn.execute("UPDATE contacts SET user_id=1 WHERE user_id IS NULL OR user_id=1")
        conn.commit()
    conn.close()

init_default_user()

def get_uid(request: Request) -> int:
    import time as _time
    token = request.headers.get('authorization', '')
    uid = active_sessions.get(token, 0)
    if uid and _time.time() - session_created.get(token, 0) > SESSION_TTL:
        active_sessions.pop(token, None)
        session_created.pop(token, None)
        return 0
    return uid

class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        open_paths = ('/api/auth/', '/profile/')
        if path.startswith('/api/') and not any(path.startswith(p) for p in open_paths):
            import time as _time
            _tok = request.headers.get('authorization','')
            _expired = _tok in active_sessions and _time.time() - session_created.get(_tok, 0) > SESSION_TTL
            if _expired:
                active_sessions.pop(_tok, None); session_created.pop(_tok, None)
            if _tok not in active_sessions:
                return JSONResponse(status_code=401, content={"detail": "認証が必要です"})
        return await call_next(request)

app.add_middleware(AuthMiddleware)


class ContactIn(BaseModel):
    name: str
    reading: Optional[str] = ''
    company: Optional[str] = ''
    chapter: Optional[str] = ''
    category: Optional[str] = ''
    last_meeting_date: Optional[str] = ''
    business_description: Optional[str] = ''
    area: Optional[str] = ''
    birthplace: Optional[str] = ''
    residence: Optional[str] = ''
    spouse: Optional[str] = ''
    family: Optional[str] = ''
    previous_jobs: Optional[str] = ''
    hobbies: Optional[str] = ''
    experience_years: Optional[str] = ''
    success_key: Optional[str] = ''
    selling_points: Optional[str] = ''
    target_customers: Optional[str] = ''
    referral_intro: Optional[str] = ''
    request: Optional[str] = ''
    goals: Optional[str] = ''
    accomplishments: Optional[str] = ''
    interests: Optional[str] = ''
    networks: Optional[str] = ''
    skills: Optional[str] = ''
    introduction: Optional[str] = ''


class MemoIn(BaseModel):
    content: str


class ReminderIn(BaseModel):
    remind_date: str
    message: Optional[str] = ''


class ReferralIn(BaseModel):
    direction: str
    date: Optional[str] = ''
    description: Optional[str] = ''
    result: Optional[str] = '進行中'
    amount: Optional[int] = 0


# ── Contacts ──────────────────────────────────────────────
@app.get("/api/contacts")
def list_contacts(request: Request, q: str = ''):
    conn = get_db()
    uid = get_uid(request)
    if q:
        like = f'%{q}%'
        rows = conn.execute(
            "SELECT * FROM contacts WHERE user_id=? AND (name LIKE ? OR company LIKE ? OR category LIKE ? OR chapter LIKE ?) ORDER BY updated_at DESC",
            (uid, like, like, like, like)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM contacts WHERE user_id=? ORDER BY updated_at DESC", (uid,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/contacts", status_code=201)
def create_contact(request: Request, c: ContactIn):
    uid = get_uid(request)
    conn = get_db()
    cur = conn.execute(
        """INSERT INTO contacts (user_id,name,reading,company,chapter,category,last_meeting_date,
           business_description,area,birthplace,residence,spouse,family,previous_jobs,hobbies,
           experience_years,success_key,selling_points,target_customers,referral_intro,request,
           goals,accomplishments,interests,networks,skills,introduction)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (uid,c.name,c.reading,c.company,c.chapter,c.category,c.last_meeting_date,
         c.business_description,c.area,c.birthplace,c.residence,c.spouse,c.family,
         c.previous_jobs,c.hobbies,c.experience_years,c.success_key,c.selling_points,
         c.target_customers,c.referral_intro,c.request,c.goals,c.accomplishments,
         c.interests,c.networks,c.skills,c.introduction)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM contacts WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return dict(row)


@app.get("/api/contacts/{cid}")
def get_contact(request: Request, cid: int):
    uid = get_uid(request)
    conn = get_db()
    row = conn.execute("SELECT * FROM contacts WHERE id=? AND user_id=?", (cid, uid)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404)
    return dict(row)


@app.put("/api/contacts/{cid}")
def update_contact(request: Request, cid: int, c: ContactIn):
    uid = get_uid(request)
    conn = get_db()
    conn.execute(
        """UPDATE contacts SET name=?,reading=?,company=?,chapter=?,category=?,last_meeting_date=?,
           business_description=?,area=?,birthplace=?,residence=?,spouse=?,family=?,previous_jobs=?,
           hobbies=?,experience_years=?,success_key=?,selling_points=?,target_customers=?,
           referral_intro=?,request=?,goals=?,accomplishments=?,interests=?,networks=?,skills=?,
           introduction=?,updated_at=datetime('now','localtime') WHERE id=? AND user_id=?""",
        (c.name,c.reading,c.company,c.chapter,c.category,c.last_meeting_date,
         c.business_description,c.area,c.birthplace,c.residence,c.spouse,c.family,
         c.previous_jobs,c.hobbies,c.experience_years,c.success_key,c.selling_points,
         c.target_customers,c.referral_intro,c.request,c.goals,c.accomplishments,
         c.interests,c.networks,c.skills,c.introduction,cid,uid)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM contacts WHERE id=?", (cid,)).fetchone()
    conn.close()
    return dict(row)


@app.delete("/api/contacts/{cid}")
def delete_contact(request: Request, cid: int):
    uid = get_uid(request)
    conn = get_db()
    conn.execute("DELETE FROM contacts WHERE id=? AND user_id=?", (cid, uid))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── Memos ─────────────────────────────────────────────────
@app.get("/api/contacts/{cid}/memos")
def list_memos(request: Request, cid: int):
    uid = get_uid(request)
    conn = get_db()
    if not conn.execute("SELECT id FROM contacts WHERE id=? AND user_id=?", (cid, uid)).fetchone():
        conn.close(); raise HTTPException(404)
    rows = conn.execute("SELECT * FROM memos WHERE contact_id=? ORDER BY created_at DESC", (cid,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/contacts/{cid}/memos", status_code=201)
def create_memo(request: Request, cid: int, m: MemoIn):
    uid = get_uid(request)
    conn = get_db()
    if not conn.execute("SELECT id FROM contacts WHERE id=? AND user_id=?", (cid, uid)).fetchone():
        conn.close(); raise HTTPException(404)
    cur = conn.execute("INSERT INTO memos (contact_id,content) VALUES (?,?)", (cid, m.content))
    conn.commit()
    row = conn.execute("SELECT * FROM memos WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return dict(row)


@app.delete("/api/memos/{mid}")
def delete_memo(request: Request, mid: int):
    uid = get_uid(request)
    conn = get_db()
    conn.execute("DELETE FROM memos WHERE id=? AND contact_id IN (SELECT id FROM contacts WHERE user_id=?)", (mid, uid))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── Reminders ─────────────────────────────────────────────
@app.get("/api/contacts/{cid}/reminders")
def list_reminders(request: Request, cid: int):
    uid = get_uid(request)
    conn = get_db()
    if not conn.execute("SELECT id FROM contacts WHERE id=? AND user_id=?", (cid, uid)).fetchone():
        conn.close(); raise HTTPException(404)
    rows = conn.execute("SELECT * FROM reminders WHERE contact_id=? ORDER BY remind_date", (cid,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/contacts/{cid}/reminders", status_code=201)
async def create_reminder(request: Request, cid: int, r: ReminderIn):
    uid = get_uid(request)
    conn = get_db()
    contact = conn.execute("SELECT name FROM contacts WHERE id=?", (cid,)).fetchone()
    cur = conn.execute("INSERT INTO reminders (contact_id,remind_date,message) VALUES (?,?,?)",
                       (cid, r.remind_date, r.message))
    conn.commit()
    row = conn.execute("SELECT * FROM reminders WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    result = dict(row)
    if contact and uid:
        event_url = await create_google_calendar_event(uid, f"1on1: {contact['name']}", r.remind_date, r.message or '')
        if event_url:
            result['google_event_url'] = event_url
    return result


@app.put("/api/reminders/{rid}/toggle")
def toggle_reminder(request: Request, rid: int):
    uid = get_uid(request)
    conn = get_db()
    conn.execute("UPDATE reminders SET done = 1 - done WHERE id=? AND contact_id IN (SELECT id FROM contacts WHERE user_id=?)", (rid, uid))
    conn.commit()
    row = conn.execute("SELECT * FROM reminders WHERE id=?", (rid,)).fetchone()
    conn.close()
    return dict(row)


@app.delete("/api/reminders/{rid}")
def delete_reminder(request: Request, rid: int):
    uid = get_uid(request)
    conn = get_db()
    conn.execute("DELETE FROM reminders WHERE id=? AND contact_id IN (SELECT id FROM contacts WHERE user_id=?)", (rid, uid))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── 認証エンドポイント ────────────────────────────────────

class LoginIn(BaseModel):
    username: str  # メールアドレスまたは旧ユーザー名（後方互換）
    password: str

class RegisterIn(BaseModel):
    email: str
    display_name: str
    password: str

class ChangePwIn(BaseModel):
    current_password: str
    new_password: str

class ChangeEmailIn(BaseModel):
    new_email: str

def _valid_email(e: str) -> bool:
    return bool(re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', e))

@app.post("/api/auth/login")
def login(data: LoginIn):
    conn = get_db()
    identifier = data.username.strip()
    # メールアドレスで検索（優先）→ 旧ユーザー名でフォールバック
    row = conn.execute("SELECT id,pw_hash,pw_salt,display_name,username FROM users WHERE email=?", (identifier,)).fetchone()
    if not row:
        row = conn.execute("SELECT id,pw_hash,pw_salt,display_name,username FROM users WHERE username=?", (identifier,)).fetchone()
    conn.close()
    if not row or not verify_pw(data.password, row[1], row[2]):
        raise HTTPException(401, detail="メールアドレスまたはパスワードが違います")
    token = secrets.token_hex(32)
    active_sessions[token] = row[0]
    session_created[token] = __import__("time").time()
    return {"token": token, "display_name": row[3], "username": row[4]}

@app.post("/api/auth/register")
def register(data: RegisterIn):
    email = data.email.strip().lower()
    if not _valid_email(email):
        raise HTTPException(400, detail="有効なメールアドレスを入力してください")
    if len(data.password) < 6:
        raise HTTPException(400, detail="パスワードは6文字以上にしてください")
    h, s = hash_pw(data.password)
    conn = get_db()
    try:
        # usernameもemailと同じ値にしてSSO統合しやすくする
        conn.execute("INSERT INTO users (username,display_name,email,pw_hash,pw_salt,auth_type) VALUES (?,?,?,?,?,'password')",
                     (email, data.display_name.strip(), email, h, s))
        conn.commit()
    except sqlite3.IntegrityError:
        raise HTTPException(400, detail="そのメールアドレスは既に登録されています")
    finally:
        conn.close()
    return {"ok": True}

@app.put("/api/auth/email")
def change_email(request: Request, data: ChangeEmailIn):
    uid = get_uid(request)
    new_email = data.new_email.strip().lower()
    if not _valid_email(new_email):
        raise HTTPException(400, detail="有効なメールアドレスを入力してください")
    conn = get_db()
    existing = conn.execute("SELECT id FROM users WHERE email=? AND id!=?", (new_email, uid)).fetchone()
    if existing:
        raise HTTPException(400, detail="そのメールアドレスは既に使われています")
    conn.execute("UPDATE users SET email=? WHERE id=?", (new_email, uid))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.get("/api/auth/sso")
def sso_login(token: str):
    secret = os.environ.get("BNI_SSO_SECRET", "")
    if not secret:
        raise HTTPException(500, detail="SSO not configured")
    try:
        payload_b64, sig = token.rsplit('.', 1)
        padding = 4 - len(payload_b64) % 4
        payload_str = base64.urlsafe_b64decode(payload_b64 + ('=' * (padding % 4))).decode()
        expected_sig = hmac.new(secret.encode(), payload_str.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(sig, expected_sig):
            raise HTTPException(400, detail="invalid token")
        payload = json.loads(payload_str)
        if payload.get('exp', 0) < time.time() * 1000:
            raise HTTPException(400, detail="token expired")
        # sigを一意キーとしてリプレイ攻撃を防止（NiceMeetトークンにjtiなし）
        if sig in used_sso_tokens:
            raise HTTPException(400, detail="invalid token")
        used_sso_tokens.add(sig)
        name = payload.get('name', '').strip()
        email = payload.get('email', '').strip()
        if not name:
            raise HTTPException(400, detail="name missing")
        conn = get_db()
        # メールアドレス優先で既存ユーザーを検索（スタンドアロン→NiceMeet統合対応）
        row = None
        if email:
            row = conn.execute("SELECT id, display_name, username FROM users WHERE email=?", (email,)).fetchone()
        if not row:
            row = conn.execute("SELECT id, display_name, username FROM users WHERE username=?", (name,)).fetchone()
        if row:
            # 既存ユーザーをSSO認証タイプに昇格
            conn.execute("UPDATE users SET auth_type='sso', email=? WHERE id=?", (email, row[0]))
            conn.commit()
            user_id, display_name, username = row[0], row[1] or name, row[2]
        else:
            # 新規ユーザー作成（SSO専用、ランダムパスワード）
            h, s_pw = hash_pw(secrets.token_hex(16))
            conn.execute(
                "INSERT INTO users (username, display_name, email, pw_hash, pw_salt, auth_type) VALUES (?,?,?,?,?,'sso')",
                (name, name, email, h, s_pw)
            )
            conn.commit()
            row = conn.execute("SELECT id FROM users WHERE username=?", (name,)).fetchone()
            user_id, display_name, username = row[0], name, name
        conn.close()
        import time as _time
        session_token = secrets.token_hex(32)
        active_sessions[session_token] = user_id
        session_created[session_token] = _time.time()
        return {"token": session_token, "display_name": display_name, "username": username}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, detail=str(e))

@app.post("/api/auth/logout")
def logout(request: Request):
    active_sessions.pop(request.headers.get('authorization',''), None)
    return {"ok": True}

@app.get("/api/me")
def get_me(request: Request):
    uid = get_uid(request)
    conn = get_db()
    row = conn.execute(
        "SELECT username, display_name, email, auth_type, plan, stripe_customer_id FROM users WHERE id=?", (uid,)
    ).fetchone()
    conn.close()
    if not row:
        raise HTTPException(401, detail="unauthorized")
    return {
        "username": row[0], "display_name": row[1], "email": row[2] or '',
        "auth_type": row[3] or 'password', "plan": row[4] or 'free',
        "has_stripe": bool(row[5])
    }

@app.post("/api/stripe/checkout")
async def stripe_checkout(request: Request):
    uid = get_uid(request)
    conn = get_db()
    row = conn.execute("SELECT email, display_name, stripe_customer_id FROM users WHERE id=?", (uid,)).fetchone()
    if not row:
        raise HTTPException(401, detail="unauthorized")
    email, display_name, customer_id = row[0], row[1], row[2]
    try:
        if not customer_id:
            customer = stripe.Customer.create(email=email or None, name=display_name or None,
                                              metadata={"bni_user_id": str(uid)})
            customer_id = customer.id
            conn.execute("UPDATE users SET stripe_customer_id=? WHERE id=?", (customer_id, uid))
            conn.commit()
        conn.close()
        session = stripe.checkout.Session.create(
            customer=customer_id,
            payment_method_types=["card"],
            line_items=[{"price": os.environ.get("STRIPE_PRICE_ID",""), "quantity": 1}],
            mode="subscription",
            subscription_data={"trial_period_days": 30},
            payment_method_collection="always",
            success_url="https://gaiaarts.org/bni/?plan=success",
            cancel_url="https://gaiaarts.org/bni/",
            locale="ja",
        )
        return {"url": session.url}
    except Exception as e:
        raise HTTPException(500, detail=str(e))

@app.post("/api/stripe/webhook")
async def stripe_webhook(request: Request):
    payload = await request.body()
    sig = request.headers.get("stripe-signature", "")
    webhook_secret = os.environ.get("STRIPE_WEBHOOK_SECRET", "")
    try:
        if webhook_secret:
            event = stripe.Webhook.construct_event(payload, sig, webhook_secret)
        else:
            event = json.loads(payload)
    except Exception as e:
        raise HTTPException(400, detail=str(e))
    etype = event.get("type", "")
    obj = event["data"]["object"]
    customer_id = obj.get("customer")
    if not customer_id:
        return {"ok": True}
    conn = get_db()
    if etype == "customer.subscription.deleted" or etype == "invoice.payment_failed":
        conn.execute("UPDATE users SET plan='free' WHERE stripe_customer_id=?", (customer_id,))
    elif etype in ("customer.subscription.created", "customer.subscription.updated", "invoice.payment_succeeded"):
        status = obj.get("status", "")
        if status in ("active", "trialing"):
            conn.execute("UPDATE users SET plan='paid' WHERE stripe_customer_id=?", (customer_id,))
        elif status in ("canceled", "unpaid", "past_due"):
            conn.execute("UPDATE users SET plan='free' WHERE stripe_customer_id=?", (customer_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.get("/api/stripe/portal")
async def stripe_portal(request: Request):
    uid = get_uid(request)
    conn = get_db()
    row = conn.execute("SELECT stripe_customer_id FROM users WHERE id=?", (uid,)).fetchone()
    conn.close()
    if not row or not row[0]:
        raise HTTPException(400, detail="no subscription")
    try:
        session = stripe.billing_portal.Session.create(
            customer=row[0],
            return_url="https://gaiaarts.org/bni/",
        )
        return {"url": session.url}
    except Exception as e:
        raise HTTPException(500, detail=str(e))

@app.post("/api/auth/change-password")
def change_password(request: Request, data: ChangePwIn):
    uid = get_uid(request)
    conn = get_db()
    row = conn.execute("SELECT pw_hash,pw_salt FROM users WHERE id=?", (uid,)).fetchone()
    conn.close()
    if not row or not verify_pw(data.current_password, row[0], row[1]):
        raise HTTPException(400, detail="現在のパスワードが違います")
    new_h, new_s = hash_pw(data.new_password)
    conn = get_db()
    conn.execute("UPDATE users SET pw_hash=?,pw_salt=? WHERE id=?", (new_h, new_s, uid))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── ダッシュボード ─────────────────────────────────────────

@app.get("/api/dashboard")
def dashboard(request: Request):
    uid = get_uid(request)
    conn = get_db()
    total = conn.execute("SELECT COUNT(*) FROM contacts WHERE user_id=?", (uid,)).fetchone()[0]
    reminders = conn.execute(
        "SELECT COUNT(*) FROM reminders r JOIN contacts c ON c.id=r.contact_id WHERE c.user_id=? AND r.done=0", (uid,)
    ).fetchone()[0]
    chapters = conn.execute("SELECT COUNT(DISTINCT chapter) FROM contacts WHERE user_id=? AND chapter!=''", (uid,)).fetchone()[0]
    monthly = conn.execute("""
        SELECT substr(last_meeting_date,1,7) as m, COUNT(*) as cnt
        FROM contacts WHERE user_id=? AND last_meeting_date!='' AND last_meeting_date IS NOT NULL
        GROUP BY m ORDER BY m DESC LIMIT 12
    """, (uid,)).fetchall()
    upcoming = conn.execute("""
        SELECT r.remind_date, r.message, c.name FROM reminders r
        JOIN contacts c ON c.id=r.contact_id
        WHERE c.user_id=? AND r.done=0 ORDER BY r.remind_date LIMIT 5
    """, (uid,)).fetchall()
    conn.close()
    return {
        "total_contacts": total,
        "pending_reminders": reminders,
        "chapters": chapters,
        "monthly": [{"month": r[0], "count": r[1]} for r in reversed(list(monthly))],
        "upcoming_reminders": [{"date": r[0], "message": r[1], "name": r[2]} for r in upcoming]
    }


# ── マイプロフィール ───────────────────────────────────────

@app.get("/api/my-profile")
def get_my_profile(request: Request):
    uid = get_uid(request)
    conn = get_db()
    row = conn.execute("SELECT profile_data, username, display_name FROM users WHERE id=?", (uid,)).fetchone()
    conn.close()
    if not row: return {}
    d = json.loads(row[0] or '{}')
    d['username'] = row[1]
    d['display_name'] = row[2]
    return d

@app.put("/api/my-profile")
async def update_my_profile(request: Request):
    uid = get_uid(request)
    data = await request.json()
    display_name = data.get('display_name', '')
    profile_json = json.dumps(data, ensure_ascii=False)
    conn = get_db()
    conn.execute("UPDATE users SET profile_data=?, display_name=? WHERE id=?", (profile_json, display_name, uid))
    conn.commit()
    conn.close()
    return data

@app.get("/api/public-profile")
def public_profile_data():
    data = get_setting('my_profile_data')
    if not data:
        raise HTTPException(404)
    p = json.loads(data)
    if not p.get('public', False):
        raise HTTPException(404, detail="プロフィールは非公開です")
    return p

@app.get("/profile/{slug}")
def public_profile_page(slug: str):
    stored_slug = get_setting('profile_slug') or 'my-profile'
    if slug != stored_slug:
        raise HTTPException(404)
    data = get_setting('my_profile_data')
    if not data:
        raise HTTPException(404)
    p = json.loads(data)
    if not p.get('public', False):
        return HTMLResponse("<h2>このプロフィールは非公開です</h2>", status_code=403)
    name = _html.escape(p.get('name',''))
    category = _html.escape(p.get('category',''))
    company = _html.escape(p.get('company',''))
    chapter = _html.escape(p.get('chapter',''))
    business = _html.escape(p.get('business_description','')).replace('\n','<br>')
    selling = _html.escape(p.get('selling_points','')).replace('\n','<br>')
    target = _html.escape(p.get('target_customers','')).replace('\n','<br>')
    referral = _html.escape(p.get('referral_intro','')).replace('\n','<br>')
    html = f"""<!DOCTYPE html>
<html lang="ja"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{name} | BNI プロフィール</title>
<script src="https://cdn.tailwindcss.com"></script></head>
<body class="bg-gray-50 min-h-screen">
<div class="max-w-2xl mx-auto py-10 px-4">
  <div class="bg-white rounded-2xl shadow p-8">
    <h1 class="text-2xl font-bold text-gray-900">{name}</h1>
    <div class="text-blue-600 font-medium mt-1">{category}</div>
    <div class="text-gray-500 text-sm mt-0.5">{company}{' / ' + chapter if chapter else ''}</div>
    {'<hr class="my-5"><h2 class="font-bold text-gray-700 mb-2">事業内容</h2><p class="text-sm text-gray-800">' + business + '</p>' if business else ''}
    {'<hr class="my-5"><h2 class="font-bold text-gray-700 mb-2">独自の強み</h2><p class="text-sm text-gray-800">' + selling + '</p>' if selling else ''}
    {'<hr class="my-5"><h2 class="font-bold text-gray-700 mb-2">こんな方を探しています</h2><p class="text-sm text-gray-800">' + target + '</p>' if target else ''}
    {'<hr class="my-5"><h2 class="font-bold text-gray-700 mb-2">紹介の切り出し方</h2><p class="text-sm text-gray-800">' + referral + '</p>' if referral else ''}
  </div>
</div></body></html>"""
    return HTMLResponse(html)


# ── AI 紹介文生成 ─────────────────────────────────────────
@app.post("/api/contacts/{cid}/generate-introduction")
def generate_introduction(cid: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM contacts WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404)
    c = dict(row)

    openai_env = Path.home() / '.secrets' / 'openai.env'
    if openai_env.exists():
        for line in openai_env.read_text().splitlines():
            if line.startswith('OPENAI_API_KEY='):
                os.environ['OPENAI_API_KEY'] = line.split('=', 1)[1].strip('"\'')

    try:
        from openai import OpenAI
        client = OpenAI()
        prompt = f"""以下のBNIメンバーの情報をもとに、他のメンバーが紹介するための自然な紹介文を200字程度で作成してください。

名前: {c.get('name','')}
職種: {c.get('category','')}
会社名: {c.get('company','')}
事業内容: {c.get('business_description','')}
独自の強み: {c.get('selling_points','')}
ターゲット顧客: {c.get('target_customers','')}
紹介の切り出し方: {c.get('referral_intro','')}

BNIメンバーが実際に使える、自然で簡潔な紹介文を書いてください。"""

        res = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=400
        )
        text = res.choices[0].message.content.strip()

        db = get_db()
        db.execute("UPDATE contacts SET introduction=?,updated_at=datetime('now','localtime') WHERE id=?", (text, cid))
        db.commit()
        db.close()
        return {"introduction": text}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ── Googleカレンダー連携 ───────────────────────────────────

def load_google_credentials():
    google_env = Path.home() / '.secrets' / 'google.env'
    if google_env.exists():
        for line in google_env.read_text().splitlines():
            if '=' in line and not line.startswith('#'):
                key, val = line.split('=', 1)
                os.environ[key.strip()] = val.strip().strip('"\'')


async def create_google_calendar_event(uid: int, title: str, date: str, description: str = '') -> str | None:
    conn = get_db()
    row = conn.execute("SELECT access_token, refresh_token FROM google_tokens WHERE user_id=?", (uid,)).fetchone()
    conn.close()
    if not row:
        return None

    load_google_credentials()
    client_id = os.environ.get('GOOGLE_CLIENT_ID', '')
    client_secret = os.environ.get('GOOGLE_CLIENT_SECRET', '')
    event_body = {"summary": title, "description": description,
                  "start": {"date": date}, "end": {"date": date}}

    async with httpx.AsyncClient() as client:
        r = await client.post(
            'https://www.googleapis.com/calendar/v3/calendars/primary/events',
            json=event_body,
            headers={'Authorization': f'Bearer {row["access_token"]}'}
        )
        if r.status_code == 401:
            ref = await client.post('https://oauth2.googleapis.com/token', data={
                'refresh_token': row['refresh_token'],
                'client_id': client_id,
                'client_secret': client_secret,
                'grant_type': 'refresh_token'
            })
            if ref.status_code != 200:
                return None
            new_token = ref.json().get('access_token')
            conn = get_db()
            conn.execute("UPDATE google_tokens SET access_token=? WHERE user_id=?", (new_token, uid))
            conn.commit()
            conn.close()
            r = await client.post(
                'https://www.googleapis.com/calendar/v3/calendars/primary/events',
                json=event_body,
                headers={'Authorization': f'Bearer {new_token}'}
            )
        if r.status_code in (200, 201):
            return r.json().get('htmlLink')
    return None


@app.get("/api/auth/google/calendar")
def google_calendar_connect(request: Request, token: str = ''):
    uid = active_sessions.get(token) or get_uid(request)
    if not uid:
        raise HTTPException(401)
    load_google_credentials()
    client_id = os.environ.get('GOOGLE_CLIENT_ID', '')
    if not client_id:
        raise HTTPException(500, detail="Google認証が未設定です")
    state = secrets.token_hex(16)
    oauth_states[state] = uid
    params = {
        'client_id': client_id,
        'redirect_uri': GOOGLE_REDIRECT_URI,
        'response_type': 'code',
        'scope': 'https://www.googleapis.com/auth/calendar.events',
        'access_type': 'offline',
        'prompt': 'consent',
        'state': state
    }
    return RedirectResponse('https://accounts.google.com/o/oauth2/v2/auth?' + urllib.parse.urlencode(params))


@app.get("/api/auth/google/callback")
async def google_calendar_callback(code: str = None, state: str = None, error: str = None):
    base_url = "https://gaiaarts.org/bni/"
    if error or not code or not state:
        return RedirectResponse(base_url + "?google_error=1")
    uid = oauth_states.pop(state, None)
    if not uid:
        return RedirectResponse(base_url + "?google_error=1")
    load_google_credentials()
    async with httpx.AsyncClient() as client:
        r = await client.post('https://oauth2.googleapis.com/token', data={
            'code': code,
            'client_id': os.environ.get('GOOGLE_CLIENT_ID', ''),
            'client_secret': os.environ.get('GOOGLE_CLIENT_SECRET', ''),
            'redirect_uri': GOOGLE_REDIRECT_URI,
            'grant_type': 'authorization_code'
        })
    if r.status_code != 200:
        return RedirectResponse(base_url + "?google_error=1")
    tokens = r.json()
    conn = get_db()
    conn.execute(
        "INSERT OR REPLACE INTO google_tokens (user_id, access_token, refresh_token) VALUES (?,?,?)",
        (uid, tokens.get('access_token', ''), tokens.get('refresh_token', ''))
    )
    conn.commit()
    conn.close()
    return RedirectResponse(base_url + "?google_connected=1")


@app.get("/api/auth/google/status")
def google_calendar_status(request: Request):
    uid = get_uid(request)
    conn = get_db()
    row = conn.execute("SELECT connected_at FROM google_tokens WHERE user_id=?", (uid,)).fetchone()
    conn.close()
    return {"connected": row is not None, "connected_at": row['connected_at'] if row else None}


@app.delete("/api/auth/google/calendar")
def google_calendar_disconnect(request: Request):
    uid = get_uid(request)
    conn = get_db()
    conn.execute("DELETE FROM google_tokens WHERE user_id=?", (uid,))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.post("/api/reminders/{rid}/sync-google")
async def sync_reminder_google(request: Request, rid: int):
    uid = get_uid(request)
    conn = get_db()
    row = conn.execute(
        "SELECT r.*, c.name as contact_name FROM reminders r JOIN contacts c ON c.id=r.contact_id WHERE r.id=?",
        (rid,)
    ).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404)
    url = await create_google_calendar_event(uid, f"1on1: {row['contact_name']}", row['remind_date'], row['message'] or '')
    if not url:
        raise HTTPException(400, detail="Googleカレンダーへの同期に失敗しました。連携を確認してください。")
    return {"event_url": url}


# ── Referrals ─────────────────────────────────────────────
@app.get("/api/contacts/{cid}/referrals")
def list_referrals(request: Request, cid: int):
    uid = get_uid(request)
    conn = get_db()
    if not conn.execute("SELECT id FROM contacts WHERE id=? AND user_id=?", (cid, uid)).fetchone():
        conn.close(); raise HTTPException(404)
    rows = conn.execute(
        "SELECT * FROM referrals WHERE contact_id=? ORDER BY date DESC, created_at DESC", (cid,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/contacts/{cid}/referrals", status_code=201)
def create_referral(request: Request, cid: int, r: ReferralIn):
    uid = get_uid(request)
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO referrals (user_id,contact_id,direction,date,description,result,amount) VALUES (?,?,?,?,?,?,?)",
        (uid, cid, r.direction, r.date, r.description, r.result, r.amount)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM referrals WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return dict(row)


@app.put("/api/referrals/{rid}")
def update_referral(request: Request, rid: int, r: ReferralIn):
    uid = get_uid(request)
    conn = get_db()
    conn.execute(
        "UPDATE referrals SET direction=?,date=?,description=?,result=?,amount=? WHERE id=? AND user_id=?",
        (r.direction, r.date, r.description, r.result, r.amount, rid, uid)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM referrals WHERE id=?", (rid,)).fetchone()
    conn.close()
    return dict(row)


@app.delete("/api/referrals/{rid}")
def delete_referral(request: Request, rid: int):
    uid = get_uid(request)
    conn = get_db()
    conn.execute("DELETE FROM referrals WHERE id=? AND user_id=?", (rid, uid))
    conn.commit()
    conn.close()
    return {"ok": True}


# ── カレンダー ────────────────────────────────────────────
@app.get("/api/calendar/google")
async def get_google_calendar_events(request: Request, year: int, month: int):
    uid = get_uid(request)
    conn = get_db()
    row = conn.execute("SELECT access_token, refresh_token FROM google_tokens WHERE user_id=?", (uid,)).fetchone()
    conn.close()
    if not row:
        return {"events": [], "connected": False}

    load_google_credentials()
    time_min = f"{year}-{month:02d}-01T00:00:00Z"
    next_y, next_m = (year + 1, 1) if month == 12 else (year, month + 1)
    time_max = f"{next_y}-{next_m:02d}-01T00:00:00Z"

    async def fetch_events(token):
        async with httpx.AsyncClient() as client:
            return await client.get(
                'https://www.googleapis.com/calendar/v3/calendars/primary/events',
                params={'timeMin': time_min, 'timeMax': time_max,
                        'singleEvents': 'true', 'orderBy': 'startTime', 'maxResults': 200},
                headers={'Authorization': f'Bearer {token}'}
            )

    r = await fetch_events(row['access_token'])
    if r.status_code == 401:
        async with httpx.AsyncClient() as client:
            ref = await client.post('https://oauth2.googleapis.com/token', data={
                'refresh_token': row['refresh_token'],
                'client_id': os.environ.get('GOOGLE_CLIENT_ID', ''),
                'client_secret': os.environ.get('GOOGLE_CLIENT_SECRET', ''),
                'grant_type': 'refresh_token'
            })
        if ref.status_code != 200:
            return {"events": [], "connected": True, "error": "トークンの更新に失敗しました"}
        new_token = ref.json().get('access_token')
        conn = get_db()
        conn.execute("UPDATE google_tokens SET access_token=? WHERE user_id=?", (new_token, uid))
        conn.commit()
        conn.close()
        r = await fetch_events(new_token)

    if r.status_code != 200:
        return {"events": [], "connected": True, "error": "Googleカレンダーの取得に失敗しました"}

    events = []
    for item in r.json().get('items', []):
        s = item.get('start', {})
        date = s.get('date') or (s.get('dateTime', '')[:10])
        events.append({
            'id': item.get('id', ''),
            'title': item.get('summary', '(タイトルなし)'),
            'date': date,
            'url': item.get('htmlLink', ''),
            'time': s.get('dateTime', '')
        })
    return {"events": events, "connected": True}


@app.get("/api/calendar")
def get_calendar(request: Request, year: int, month: int):
    uid = get_uid(request)
    month_str = f"{year}-{month:02d}"
    conn = get_db()
    reminders = conn.execute("""
        SELECT r.id, r.remind_date, r.message, r.done, c.name as contact_name, c.id as contact_id
        FROM reminders r JOIN contacts c ON c.id=r.contact_id
        WHERE c.user_id=? AND r.remind_date LIKE ?
        ORDER BY r.remind_date
    """, (uid, f"{month_str}%")).fetchall()
    meetings = conn.execute("""
        SELECT id, name, category, company, last_meeting_date
        FROM contacts WHERE user_id=? AND last_meeting_date LIKE ?
        ORDER BY last_meeting_date
    """, (uid, f"{month_str}%")).fetchall()
    conn.close()
    return {
        "reminders": [dict(r) for r in reminders],
        "meetings": [dict(m) for m in meetings]
    }


# ── 1on1優先度サジェスト ───────────────────────────────────
@app.get("/api/suggestions/next-meetings")
def suggest_next_meetings(request: Request):
    uid = get_uid(request)
    conn = get_db()
    contacts = conn.execute("SELECT * FROM contacts WHERE user_id=?", (uid,)).fetchall()
    today = datetime.date.today()
    results = []

    for c in [dict(r) for r in contacts]:
        days_since = None
        if c.get('last_meeting_date'):
            try:
                last = datetime.date.fromisoformat(c['last_meeting_date'])
                days_since = (today - last).days
            except Exception:
                pass

        ref_count = conn.execute(
            "SELECT COUNT(*) FROM referrals WHERE contact_id=?", (c['id'],)
        ).fetchone()[0]

        overdue = conn.execute(
            "SELECT COUNT(*) FROM reminders WHERE contact_id=? AND done=0 AND remind_date<=?",
            (c['id'], str(today))
        ).fetchone()[0]

        if days_since is None:
            day_score, reason = 60, "まだ1on1をしていません"
        elif days_since >= 180:
            day_score, reason = 50, f"{days_since}日間1on1をしていません"
        elif days_since >= 90:
            day_score, reason = 35, f"{days_since}日ぶりの1on1が必要です"
        elif days_since >= 30:
            day_score, reason = 20, f"最終1on1から{days_since}日経過"
        else:
            day_score, reason = max(0, days_since // 2), f"{days_since}日前に会いました"

        score = day_score + min(ref_count * 5, 20) + overdue * 15

        tags = []
        if days_since is None:
            tags.append("未1on1")
        elif days_since >= 90:
            tags.append(f"{days_since}日経過")
        if overdue:
            tags.append(f"期限切れリマインダー×{overdue}")
        if ref_count:
            tags.append(f"紹介{ref_count}件")

        results.append({
            "contact": {"id": c['id'], "name": c['name'], "category": c.get('category',''), "company": c.get('company','')},
            "score": score,
            "reason": reason,
            "tags": tags,
            "days_since": days_since,
            "ref_count": ref_count
        })

    conn.close()
    results.sort(key=lambda x: x['score'], reverse=True)
    return results[:10]


# ── AIマッチング ──────────────────────────────────────────
@app.post("/api/contacts/{cid}/match")
def match_contacts(request: Request, cid: int):
    uid = get_uid(request)
    conn = get_db()
    target = conn.execute("SELECT * FROM contacts WHERE id=? AND user_id=?", (cid, uid)).fetchone()
    if not target:
        raise HTTPException(404)
    others = conn.execute(
        "SELECT * FROM contacts WHERE user_id=? AND id!=? ORDER BY updated_at DESC LIMIT 30",
        (uid, cid)
    ).fetchall()
    conn.close()

    if not others:
        return {"matches": []}

    t = dict(target)
    target_summary = f"""名前: {t.get('name','')}
職種: {t.get('category','')}
事業内容: {t.get('business_description','')}
独自の強み: {t.get('selling_points','')}
ターゲット顧客: {t.get('target_customers','')}
目標: {t.get('goals','')}
スキル: {t.get('skills','')}
人脈: {t.get('networks','')}"""

    others_list = [dict(r) for r in others]
    others_summary = "\n\n".join([
        f"ID:{o['id']} 名前:{o.get('name','')} 職種:{o.get('category','')} "
        f"事業:{o.get('business_description','')[:200]} "
        f"強み:{o.get('selling_points','')[:150]} "
        f"ターゲット:{o.get('target_customers','')[:150]} "
        f"目標:{o.get('goals','')[:100]}"
        for o in others_list
    ])

    prompt = f"""あなたはBNIのビジネスマッチングAIです。
【対象者】と最も相性の良い人を【候補者リスト】から選び、上位3名をJSONで返してください。

【対象者】
{target_summary}

【候補者リスト】
{others_summary}

評価基準：
1. 紹介し合える可能性（互いのターゲット顧客・事業内容の親和性）
2. ビジネスシナジー（補完関係、協業可能性）
3. GAINS共通点（目標・興味・スキルの一致）

出力JSON:
{{"matches":[{{"contact_id":<ID>,"score":<1-100>,"reason":"<50字程度の理由>","referral_opportunity":"<具体的な紹介シナリオ>"}}]}}

上位3名のみ、JSONのみ返してください。"""

    load_openai()
    try:
        from openai import OpenAI
        client = OpenAI()
        res = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1000,
            response_format={"type": "json_object"}
        )
        data = json.loads(res.choices[0].message.content)
        matches = data.get('matches', [])

        id_map = {o['id']: o for o in others_list}
        result = []
        for m in matches:
            mid = m.get('contact_id')
            if mid in id_map:
                o = id_map[mid]
                result.append({
                    "contact": {"id": o['id'], "name": o.get('name',''), "category": o.get('category',''), "company": o.get('company','')},
                    "score": m.get('score', 0),
                    "reason": m.get('reason', ''),
                    "referral_opportunity": m.get('referral_opportunity', '')
                })
        return {"matches": result}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ── インポート ────────────────────────────────────────────

def load_openai():
    openai_env = Path.home() / '.secrets' / 'openai.env'
    if openai_env.exists():
        for line in openai_env.read_text().splitlines():
            if line.startswith('OPENAI_API_KEY='):
                os.environ['OPENAI_API_KEY'] = line.split('=', 1)[1].strip('"\'')


def extract_text_from_pdf(path: str) -> str:
    import pdfplumber
    text = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text.append(t)
    return '\n'.join(text)


def extract_text_from_docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())


def clean_csv_text(text: str) -> str:
    """CSV形式の空行・空カラムを除去して読みやすいテキストに変換"""
    import csv, io
    lines = []
    try:
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            cells = [c.strip() for c in row if c.strip()]
            if cells:
                lines.append('  '.join(cells))
    except Exception:
        # CSV解析に失敗したらそのまま返す
        return text
    return '\n'.join(lines)


def parse_with_ai(text: str) -> dict:
    load_openai()
    from openai import OpenAI
    client = OpenAI()

    prompt = f"""あなたはBNIメンバー管理システムのデータ抽出AIです。
以下のテキストはBNIの「メンバー略歴シート」または「Power 1-2-1シート」または「GAINSシート」です。
内容を解析して指定のフィールドにマッピングし、JSONで返してください。

【BNI書類の対応表】
メンバー略歴シート:
  「スピーカー：」→ name
  「チャプター」→ chapter
  「事業名：」→ company
  「専門分野：」→ category（職種・カテゴリー）
  「経験年数：」→ experience_years
  「所在地：」→ area
  「過去に経験した職業：」→ previous_jobs
  「配偶者：」→ spouse
  「その他家族：」→ family
  「出身地：」→ birthplace
  「居住地：」→ residence
  「趣味：」→ hobbies
  「私の成功の鍵は」→ success_key

Power 1-2-1シート（番号付き項目）:
  「1.氏名」「1.名前」→ name
  「2.会社名または屋号」→ company
  「３．専門分野、中心的なサービス」→ business_description（長文でもそのまま全文入れる）
  「４．他社にない強み」→ selling_points（長文でもそのまま全文入れる）
  「５．どんな人、どんな会社が良いクライアントになりますか」→ target_customers（長文でもそのまま全文入れる）
  「６．あなたについて、どう会話を切り出したらよいですか」→ referral_intro（長文でもそのまま全文入れる）

GAINSシート:
  「Goals（目標）」→ goals
  「Accomplishments（実績）」→ accomplishments
  「Interests（興味関心）」→ interests
  「Networks（人脈）」→ networks
  「Skills（スキル）」→ skills

【重要ルール】
- 長い説明文・箇条書きはそのまま全文を入れる（省略しない）
- 項目が見当たらない場合は空文字列
- 数字付き項目（１、２、３...）はPower 1-2-1の項目番号として解釈する
- 複数のシートが混在している場合はすべてのシートから抽出する

抽出フィールド（JSON keys）:
name, reading, company, chapter, category, business_description, area, birthplace, residence,
spouse, family, previous_jobs, hobbies, experience_years,
success_key, selling_points, target_customers, referral_intro, request,
goals, accomplishments, interests, networks, skills

テキスト（最大10000文字）:
{text[:10000]}

JSONのみ返してください。"""

    res = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=3000,
        response_format={"type": "json_object"}
    )
    raw = res.choices[0].message.content.strip()
    return json.loads(raw)


@app.post("/api/import/file")
async def import_file(file: UploadFile = File(...)):
    suffix = Path(file.filename).suffix.lower()
    if suffix not in ('.pdf', '.docx', '.doc', '.txt'):
        raise HTTPException(400, detail="対応形式: PDF, DOCX, TXT")

    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        if suffix == '.pdf':
            text = extract_text_from_pdf(tmp_path)
        elif suffix in ('.docx', '.doc'):
            text = extract_text_from_docx(tmp_path)
        else:
            text = Path(tmp_path).read_text(encoding='utf-8', errors='ignore')

        if not text.strip():
            raise HTTPException(400, detail="テキストを抽出できませんでした")

        if ',,' in text[:200]:
            text = clean_csv_text(text)

        parsed = parse_with_ai(text)
        return {"parsed": parsed, "raw_text": text[:1200]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, detail=str(e))
    finally:
        os.unlink(tmp_path)


def extract_text_from_xlsx(data: bytes) -> str:
    """XLSXの全シートからテキストを抽出"""
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lines = [f"=== シート: {sheet_name} ==="]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip() and str(c).strip() != 'None']
            if cells:
                sheet_lines.append('  '.join(cells))
        if len(sheet_lines) > 1:
            parts.append('\n'.join(sheet_lines))
    return '\n\n'.join(parts)


@app.post("/api/import/url")
async def import_url(url: str = Form(...)):
    export_url = url
    is_sheets = 'docs.google.com/spreadsheets' in url

    if 'docs.google.com/document' in url:
        doc_id = url.split('/d/')[1].split('/')[0]
        export_url = f'https://docs.google.com/document/d/{doc_id}/export?format=txt'
    elif is_sheets:
        sheet_id = url.split('/d/')[1].split('/')[0]
        export_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'

    try:
        async with httpx.AsyncClient(follow_redirects=True, timeout=30) as client:
            r = await client.get(export_url)
        if r.status_code != 200:
            raise HTTPException(400, detail=f"取得失敗 (HTTP {r.status_code})。ファイルが「リンクを知っている人全員が閲覧可能」に設定されているか確認してください。")

        content = r.content
        content_type = r.headers.get('content-type', '')

        if 'html' in content_type or b'accounts.google.com' in content[:500]:
            raise HTTPException(400, detail="Googleのログイン画面が返されました。ファイルの共有設定を「リンクを知っている人全員が閲覧可能」に変更してください。")

        if is_sheets or 'spreadsheet' in content_type or content[:4] == b'PK\x03\x04':
            text = extract_text_from_xlsx(content)
        else:
            text = r.text
            if ',,' in text[:200]:
                text = clean_csv_text(text)

        parsed = parse_with_ai(text)
        return {"parsed": parsed, "raw_text": text[:1200]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ── NiceMeet BNI連携 ─────────────────────────────────────
@app.post("/api/nicemeet-contact")
async def nicemeet_contact(request: Request):
    secret = request.headers.get("x-nicemeet-secret", "")
    expected = os.environ.get("NICEMEET_WEBHOOK_SECRET", "nicemeet-bni-2026")
    if secret != expected:
        raise HTTPException(403, detail="forbidden")
    data = await request.json()
    bni_user = data.get("bni_user", "")
    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    if not bni_user or not name or not email:
        raise HTTPException(400, detail="missing fields")
    is_bni = bool(data.get("is_bni_member", True))
    category = (data.get("category") or "").strip()
    chapter = (data.get("chapter") or "").strip()
    conn = get_db()
    user = conn.execute("SELECT id FROM users WHERE username=?", (bni_user,)).fetchone()
    if not user:
        conn.close()
        raise HTTPException(404, detail="user not found")
    uid = user["id"]
    existing = conn.execute(
        "SELECT id FROM contacts WHERE user_id=? AND (name=? OR introduction LIKE ?)",
        (uid, name, f"%{email}%")
    ).fetchone()
    if existing:
        conn.execute("""
            UPDATE contacts SET
                category=CASE WHEN category='' OR category IS NULL THEN ? ELSE category END,
                chapter=CASE WHEN chapter='' OR chapter IS NULL THEN ? ELSE chapter END,
                introduction=CASE WHEN introduction='' OR introduction IS NULL THEN ? ELSE introduction END,
                updated_at=datetime('now','localtime')
            WHERE id=?
        """, (category, chapter, email, existing["id"]))
        conn.commit()
        conn.close()
        return {"ok": True, "action": "updated", "id": existing["id"]}
    conn.execute("""
        INSERT INTO contacts (user_id, name, category, chapter, introduction)
        VALUES (?, ?, ?, ?, ?)
    """, (uid, name, category, chapter, email))
    conn.commit()
    new_id = conn.execute("SELECT last_insert_rowid() as id").fetchone()["id"]
    conn.close()
    return {"ok": True, "action": "created", "id": new_id}

@app.post("/api/nicemeet-webhook")
async def nicemeet_webhook(request: Request):
    secret = request.headers.get("x-nicemeet-secret", "")
    expected = os.environ.get("NICEMEET_WEBHOOK_SECRET", "nicemeet-bni-2026")
    if secret != expected:
        raise HTTPException(403, detail="forbidden")
    data = await request.json()
    conn = get_db()
    user = conn.execute("SELECT id FROM users WHERE username=?", (data.get("bni_user",""),)).fetchone()
    if not user:
        conn.close()
        raise HTTPException(404, detail="user not found")
    uid = user["id"]
    contact_id = data.get("contact_id")
    if contact_id:
        c = conn.execute("SELECT id FROM contacts WHERE id=? AND user_id=?", (contact_id, uid)).fetchone()
        if not c:
            contact_id = None
    gains = data.get("gains", {})
    conn.execute("""
        INSERT INTO one_on_ones (user_id, contact_id, contact_name, duration_minutes, transcript, summary,
            gains_goals, gains_accomplishments, gains_interests, gains_networks, gains_skills,
            referral_hints, follow_up)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (uid, contact_id, data.get("contact_name",""), data.get("duration_minutes", 0),
          data.get("transcript",""), data.get("summary",""),
          gains.get("goals",""), gains.get("accomplishments",""), gains.get("interests",""),
          gains.get("networks",""), gains.get("skills",""),
          data.get("referral_hints",""), data.get("follow_up","")))
    conn.commit()
    if contact_id and any(gains.values()):
        c = conn.execute("SELECT goals,accomplishments,interests,networks,skills FROM contacts WHERE id=?", (contact_id,)).fetchone()
        updates = {}
        for f in ["goals","accomplishments","interests","networks","skills"]:
            if not c[f] and gains.get(f):
                updates[f] = gains[f]
        if updates:
            set_clause = ", ".join(f"{k}=?" for k in updates)
            conn.execute(f"UPDATE contacts SET {set_clause}, updated_at=datetime('now','localtime') WHERE id=?",
                         list(updates.values()) + [contact_id])
            conn.commit()
    conn.close()

    # R2にも保存（GAIADrive経由）
    try:
        import datetime, urllib.request, json as _json
        drive_url = os.environ.get("DRIVE_INTERNAL_URL", "http://localhost:8309/api/internal/upload-json")
        drive_secret = os.environ.get("DRIVE_INTERNAL_SECRET", "gaia-internal-2026")
        today = datetime.date.today().isoformat()
        safe_name = (data.get("contact_name","unknown") or "unknown").replace(" ", "_")
        r2_key = f"1to1manager/one_on_ones/{today}/{data.get('bni_user','unknown')}-{safe_name}.json"
        r2_body = _json.dumps({
            "date": today,
            "bni_user": data.get("bni_user",""),
            "contact_name": data.get("contact_name",""),
            "duration_minutes": data.get("duration_minutes", 0),
            "transcript": data.get("transcript",""),
            "summary": data.get("summary",""),
            "gains": data.get("gains", {}),
            "referral_hints": data.get("referral_hints",""),
            "follow_up": data.get("follow_up","")
        }, ensure_ascii=False, indent=2)
        req = urllib.request.Request(
            drive_url,
            data=_json.dumps({"key": r2_key, "content": r2_body}).encode(),
            headers={"Content-Type": "application/json", "x-internal-secret": drive_secret},
            method="POST"
        )
        urllib.request.urlopen(req, timeout=5)
    except Exception as e:
        print(f"[R2 upload error] {e}")

    return {"ok": True}


# ── データエクスポート ────────────────────────────────────────

def build_viewer_html(contacts, one_on_ones):
    contacts_json = json.dumps(contacts, ensure_ascii=False)
    oo_json = json.dumps(one_on_ones, ensure_ascii=False)
    today = datetime.date.today().isoformat()
    return f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>1to1 Manager - ローカルアーカイブ</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#0f172a;color:#f1f5f9;height:100vh;display:flex;flex-direction:column}}
header{{background:#0f172a;border-bottom:1px solid rgba(255,255,255,0.07);padding:12px 20px;display:flex;align-items:center;gap:12px}}
header h1{{background:linear-gradient(135deg,#a5b4fc,#8b5cf6);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;font-size:1.1rem}}
.badge{{font-size:11px;background:rgba(251,191,36,0.15);color:#fbbf24;border:1px solid rgba(251,191,36,0.3);padding:3px 8px;border-radius:20px}}
.info{{margin-left:auto;font-size:12px;color:#64748b}}
.layout{{display:flex;flex:1;overflow:hidden}}
#sidebar{{width:280px;background:#1e293b;border-right:1px solid #334155;display:flex;flex-direction:column}}
.search-wrap{{padding:12px}}
#search{{width:100%;background:#0f172a;border:1px solid #334155;border-radius:8px;color:#f1f5f9;padding:8px 12px;font-size:13px;outline:none}}
#search:focus{{border-color:#6366f1}}
#search::placeholder{{color:#475569}}
#contact-list{{flex:1;overflow-y:auto}}
.c-item{{padding:12px 16px;border-bottom:1px solid #1e293b;cursor:pointer;transition:background .12s}}
.c-item:hover{{background:#334155}}
.c-item.active{{background:#1e1b4b;border-left:3px solid #6366f1}}
.c-name{{font-size:13px;font-weight:600;color:#e2e8f0}}
.c-meta{{font-size:11px;color:#64748b;margin-top:2px}}
#detail{{flex:1;overflow-y:auto;padding:24px}}
.placeholder{{display:flex;align-items:center;justify-content:center;height:100%;color:#475569;font-size:14px}}
.d-header{{margin-bottom:20px}}
.d-name{{font-size:1.3rem;font-weight:700;color:#e2e8f0;margin-bottom:4px}}
.d-meta{{font-size:12px;color:#64748b}}
.oo-card{{background:#1e293b;border:1px solid #334155;border-radius:12px;padding:16px;margin-bottom:14px}}
.oo-date{{font-size:11px;color:#64748b;margin-bottom:8px;display:flex;gap:8px;align-items:center}}
.oo-dur{{background:#1e1b4b;color:#a5b4fc;padding:2px 8px;border-radius:10px;font-size:11px}}
.oo-summary{{font-size:13px;color:#cbd5e1;line-height:1.6;margin-bottom:12px}}
.gains-grid{{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:10px}}
.gains-item{{background:#0f172a;border-radius:8px;padding:10px}}
.gains-label{{font-size:10px;color:#8b5cf6;font-weight:700;letter-spacing:.5px;margin-bottom:4px}}
.gains-val{{font-size:12px;color:#94a3b8;line-height:1.5}}
.hint-block{{background:#0f172a;border-radius:8px;padding:10px;margin-top:8px}}
.hint-label{{font-size:10px;color:#f59e0b;font-weight:700;margin-bottom:4px}}
.hint-val{{font-size:12px;color:#94a3b8;line-height:1.5}}
.transcript-toggle{{font-size:12px;color:#6366f1;cursor:pointer;margin-top:8px;display:block}}
.transcript{{font-size:11px;color:#64748b;line-height:1.7;margin-top:8px;white-space:pre-wrap;display:none;background:#0f172a;padding:10px;border-radius:8px}}
::-webkit-scrollbar{{width:4px}}::-webkit-scrollbar-thumb{{background:#334155;border-radius:4px}}
</style>
</head>
<body>
<header>
  <span style="font-size:20px">🤝</span>
  <h1>1to1 Manager</h1>
  <span class="badge">📦 ローカルアーカイブ</span>
  <span class="info">エクスポート日: {today}</span>
</header>
<div class="layout">
  <div id="sidebar">
    <div class="search-wrap">
      <input id="search" placeholder="🔍 名前・会社で検索" oninput="filterContacts(this.value)">
    </div>
    <div id="contact-list"></div>
  </div>
  <div id="detail"><div class="placeholder">← コンタクトを選択してください</div></div>
</div>
<script>
const CONTACTS = {contacts_json};
const ONE_ON_ONES = {oo_json};
let filtered = [...CONTACTS];
let selected = null;

function renderList(items) {{
  const el = document.getElementById('contact-list');
  if (!items.length) {{ el.innerHTML = '<div style="padding:20px;text-align:center;color:#475569;font-size:13px">見つかりません</div>'; return; }}
  el.innerHTML = items.map(c => {{
    const cnt = ONE_ON_ONES.filter(o => o.contact_id === c.id).length;
    return `<div class="c-item${{selected===c.id?' active':''}}" onclick="selectContact(${{c.id}})">
      <div class="c-name">${{c.name}}</div>
      <div class="c-meta">${{c.company||''}}${{c.chapter?' · '+c.chapter:''}} · 1-2-1: ${{cnt}}回</div>
    </div>`;
  }}).join('');
}}

function filterContacts(q) {{
  const lq = q.toLowerCase();
  filtered = CONTACTS.filter(c => (c.name||'').toLowerCase().includes(lq)||(c.company||'').toLowerCase().includes(lq));
  renderList(filtered);
}}

function selectContact(id) {{
  selected = id;
  renderList(filtered);
  const c = CONTACTS.find(x => x.id===id);
  const records = ONE_ON_ONES.filter(o => o.contact_id===id).sort((a,b)=>b.created_at?.localeCompare(a.created_at));
  const el = document.getElementById('detail');
  if (!c) return;
  const gainsKeys = [['goals','目標・ゴール'],['accomplishments','実績'],['interests','興味・関心'],['networks','人脈'],['skills','スキル']];
  el.innerHTML = `<div class="d-header">
    <div class="d-name">${{c.name}}</div>
    <div class="d-meta">${{c.company||''}}${{c.category?' · '+c.category:''}}${{c.chapter?' · '+c.chapter:''}}</div>
  </div>` + (records.length ? records.map((r,i) => {{
    const gains = gainsKeys.filter(([k]) => r['gains_'+k]).map(([k,label]) =>
      `<div class="gains-item"><div class="gains-label">${{label}}</div><div class="gains-val">${{r['gains_'+k]}}</div></div>`).join('');
    return `<div class="oo-card">
      <div class="oo-date"><span>${{r.created_at?.slice(0,10)||''}}</span><span class="oo-dur">${{r.duration_minutes||0}}分</span></div>
      ${{r.summary ? `<div class="oo-summary">${{r.summary}}</div>` : ''}}
      ${{gains ? `<div class="gains-grid">${{gains}}</div>` : ''}}
      ${{r.referral_hints ? `<div class="hint-block"><div class="hint-label">🔗 紹介ヒント</div><div class="hint-val">${{r.referral_hints}}</div></div>` : ''}}
      ${{r.follow_up ? `<div class="hint-block"><div class="hint-label">📌 フォローアップ</div><div class="hint-val">${{r.follow_up}}</div></div>` : ''}}
      ${{r.transcript ? `<span class="transcript-toggle" onclick="this.nextElementSibling.style.display=this.nextElementSibling.style.display==='none'?'block':'none'">📄 文字起こしを表示/非表示</span><div class="transcript" style="display:none">${{r.transcript}}</div>` : ''}}
    </div>`;
  }}).join('') : '<div style="color:#475569;font-size:13px;margin-top:20px">1-2-1の記録がありません</div>');
}}

renderList(CONTACTS);
</script>
</body>
</html>"""

@app.get("/api/export/zip")
def export_zip(request: Request):
    uid = get_uid(request)
    if not uid:
        raise HTTPException(401, detail="Unauthorized")
    conn = get_db()
    contacts = [dict(r) for r in conn.execute(
        "SELECT * FROM contacts WHERE user_id=? ORDER BY name", (uid,)).fetchall()]
    one_on_ones = [dict(r) for r in conn.execute(
        "SELECT * FROM one_on_ones WHERE user_id=? ORDER BY created_at DESC", (uid,)).fetchall()]
    conn.close()

    viewer = build_viewer_html(contacts, one_on_ones)
    today = datetime.date.today().isoformat()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('viewer.html', viewer)
        zf.writestr('data/contacts.json',
            json.dumps(contacts, ensure_ascii=False, indent=2))
        zf.writestr('data/one_on_ones.json',
            json.dumps(one_on_ones, ensure_ascii=False, indent=2))
        readme = f"""1to1 Manager データエクスポート
エクスポート日: {today}

■ ファイル構成
viewer.html          ブラウザで開くと全データを閲覧できます（サーバー不要）
data/contacts.json   コンタクト一覧（再インポート用）
data/one_on_ones.json 1-2-1記録一覧（再インポート用）

■ 使い方
1. viewer.html をブラウザで開く
2. 左のリストからコンタクトを選択
3. 右に1-2-1の記録・GAINS・文字起こしが表示されます

■ データ保持ポリシー
・ご契約中: サーバー上に全データを保持
・解約後:   90日間ダウンロード可能、その後サーバーから削除
・R2バックアップ: 3年間保持
"""
        zf.writestr('README.txt', readme)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/zip',
        headers={{'Content-Disposition': f'attachment; filename="1to1manager-{today}.zip"'}}
    )

@app.get("/api/contacts/{cid}/one-on-ones")
def get_contact_one_on_ones(cid: int, request: Request):
    uid = get_uid(request)
    if not uid:
        raise HTTPException(401)
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM one_on_ones WHERE contact_id=? AND user_id=? ORDER BY created_at DESC",
        (cid, uid)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/api/one-on-ones")
def get_all_one_on_ones(request: Request):
    uid = get_uid(request)
    if not uid:
        raise HTTPException(401)
    conn = get_db()
    rows = conn.execute("""
        SELECT o.*, COALESCE(c.name, o.contact_name) as contact_display, c.company
        FROM one_on_ones o
        LEFT JOIN contacts c ON o.contact_id = c.id
        WHERE o.user_id=?
        ORDER BY o.created_at DESC LIMIT 100
    """, (uid,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ── Static / SPA ─────────────────────────────────────────
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")


@app.get("/")
def root():
    return FileResponse(str(BASE_DIR / "static" / "index.html"))


@app.get("/lp")
@app.get("/lp.html")
def lp_page():
    return FileResponse(str(BASE_DIR / "static" / "lp.html"))


@app.get("/{path:path}")
def spa(path: str):
    return FileResponse(str(BASE_DIR / "static" / "index.html"))
